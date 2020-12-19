# coding=utf-8

import openpyxl
import numpy as np
from cleanText import cleanString
from sklearn import tree
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer

class id3Classifier:

    check=0
    model = tree.DecisionTreeClassifier(criterion="entropy")
    vectorizer = TfidfVectorizer(stop_words='english', max_df=75)
    
    
    # Get the original dataset
    def store():
        workBookOld = openpyxl.load_workbook('datasets/trainData.xlsx')
        dataSheetOld = workBookOld['trainData']

        xData = []
        yData = []

        rows = dataSheetOld.max_row

        for i in range(2, rows+1):

            if (str(dataSheetOld.cell(row = i, column = 2).value) != 'None'):
                xData.append(str(cleanString(str(dataSheetOld.cell(row = i, column = 1).value))))
                if (str(dataSheetOld.cell(row = i, column = 2).value) == "1"):
                    yData.append(1)
                else:
                    yData.append(0)

        # NOTE: to train data on the entire dataset, simply return xData and yData
        # Splitting the data like this is to obtain test cases and calculate the F-score of the learning algorithm
        xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size=0.2, random_state=0)
        return xTrain, xTest, yTrain, yTest


    # Calculating the F-score
    def calcFScore(xTest, yTest, model, vectorizer):
        
        xTestMatrix = vectorizer.transform(xTest)
        yTestMatrix = np.asarray(yTest)

        result = model.predict(xTestMatrix)
        matrix = confusion_matrix(yTestMatrix, result)

        fScore = f1_score(yTestMatrix, result, pos_label = 0)
        precision = precision_score(yTestMatrix, result, pos_label=0)
        recall = recall_score(yTestMatrix, result, pos_label=0)
        return fScore, precision, recall, matrix

    def predict(msg):
        if id3Classifier.check==0:

            # Create training data
            xTrain, xTest, yTrain, yTest = id3Classifier.store()

            id3Classifier.vectorizer = TfidfVectorizer(stop_words='english', max_df=75)

            yTrainMatrix = np.asarray(yTrain)
            xTrainMatrix = id3Classifier.vectorizer.fit_transform(xTrain)

            # Training ID3 classifier
            id3Classifier.model.fit(xTrainMatrix, yTrainMatrix)
            fScore, precision, recall, matrix = id3Classifier.calcFScore(xTest, yTest, id3Classifier.model, id3Classifier.vectorizer)
            print("fScore, precision, recall :")
            print(fScore, precision, recall)
            id3Classifier.check=1
            
        return id3Classifier.predict2(msg,id3Classifier.vectorizer,id3Classifier.model)

    # Test new data for Spam
    def predict2(emailBody,vectorizer,model):

        featureMatrix = vectorizer.transform([cleanString(emailBody)])
        result = model.predict(featureMatrix)

        if (1 in result):
            #return "Spam"
            return True
        else:
            #return "Not Spam"
            return False

'''
print(id3Classifier.predict("FreeMsg: Claim ur 250 SMS messages-Text OK to 84025 now!Use web2mobile 2 ur mates etc. Join Txt250.com for 1.50p/wk. T&C BOX139, LA32WU. 16 . Remove txtX or stop"))
print(id3Classifier.predict("FREE for 1st week! No1 Nokia tone 4 ur mob every week just txt NOKIA to 87077 Get txting and tell ur mates. zed POBox 36504 W45WQ norm150p/tone 16+"))
print(id3Classifier.predict("I have a tad issue here about the thorough refining column"))
'''


