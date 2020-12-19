# coding=utf-8

import openpyxl
import numpy as np
import random
import time
from cleanText import cleanString
from sklearn.naive_bayes import MultinomialNB
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer

class nbClassifier:

    check=0
    model = MultinomialNB()
    vectorizer = TfidfVectorizer(stop_words='english', max_df=75)
    
    
    # Get the original dataset
    def store():
        workBookOld = openpyxl.load_workbook('datasets/trainData.xlsx')
        dataSheetOld = workBookOld['trainData']

        xData = []
        yData = []

        rows = dataSheetOld.max_row

        for i in range(2, rows+1):

            if (str(dataSheetOld.cell(row = i, column = 2).value) != 'None'):
                xData.append(str(cleanString(str(dataSheetOld.cell(row = i, column = 1).value))))
                if (str(dataSheetOld.cell(row = i, column = 2).value) == "1"):
                    yData.append(1)
                else:
                    yData.append(0)

        # NOTE: to train data on the entire dataset, simply return xData and yData
        # Splitting the data like this is to obtain test cases and calculate the F-score of the learning algorithm
        xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size=0.2, random_state=0)
        return xTrain, xTest, yTrain, yTest


    # Calculating the F-score
    def calcFScore(xTest, yTest, model, vectorizer):
        
        xTestMatrix = vectorizer.transform(xTest)
        yTestMatrix = np.asarray(yTest)

        result = model.predict(xTestMatrix)
        matrix = confusion_matrix(yTestMatrix, result)

        fScore = f1_score(yTestMatrix, result, pos_label = 0)
        precision = precision_score(yTestMatrix, result, pos_label=0)
        recall = recall_score(yTestMatrix, result, pos_label=0)
        return fScore, precision, recall, matrix

    def predict(msg):
        if nbClassifier.check==0:

            # Create training data
            xTrain, xTest, yTrain, yTest = nbClassifier.store()

            nbClassifier.vectorizer = TfidfVectorizer(stop_words='english', max_df=75)
            yTrainMatrix = np.asarray(yTrain)
            xTrainMatrix = nbClassifier.vectorizer.fit_transform(xTrain)

            # Training NB classifier
            nbClassifier.model.fit(xTrainMatrix, yTrainMatrix)
            fScore, precision, recall, matrix = nbClassifier.calcFScore(xTest, yTest, nbClassifier.model, nbClassifier.vectorizer)
            print("fScore, precision, recall :")
            print(fScore, precision, recall)
            nbClassifier.check=1
            
        return nbClassifier.predict2(msg,nbClassifier.vectorizer,nbClassifier.model)

    # Test new data for Spam
    def predict2(emailBody,vectorizer,model):

        featureMatrix = vectorizer.transform([cleanString(emailBody)])
        result = model.predict(featureMatrix)
        if (1 in result):
            #return "Spam"
            return True
        else:
            #return "Not Spam"
            return False

'''
print(nbClassifier.predict("FreeMsg: Claim ur 250 SMS messages-Text OK to 84025 now!Use web2mobile 2 ur mates etc. Join Txt250.com for 1.50p/wk. T&C BOX139, LA32WU. 16 . Remove txtX or stop"))
print(nbClassifier.predict("FREE for 1st week! No1 Nokia tone 4 ur mob every week just txt NOKIA to 87077 Get txting and tell ur mates. zed POBox 36504 W45WQ norm150p/tone 16+"))
print(nbClassifier.predict("I have a tad issue here about the thorough refining column"))

'''
