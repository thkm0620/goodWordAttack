# coding=utf-8

import openpyxl
import numpy as np
from cleanText import cleanString
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer

class lrClassifier:

    check=0
    model = LogisticRegression()
    vectorizer = TfidfVectorizer(stop_words='english', max_df=75)
    
    
    # Get the original dataset
    def store():
        workBookOld = openpyxl.load_workbook('datasets/trainData.xlsx')
        dataSheetOld = workBookOld['trainData']

        xData = []
        yData = []

        rows = dataSheetOld.max_row

        for i in range(2, rows+1):

            if (str(dataSheetOld.cell(row = i, column = 2).value) != 'None'):
                xData.append(str(cleanString(str(dataSheetOld.cell(row = i, column = 1).value))))
                if (str(dataSheetOld.cell(row = i, column = 2).value) == "1"):
                    yData.append(1)
                else:
                    yData.append(0)

        # NOTE: to train data on the entire dataset, simply return xData and yData
        # Splitting the data like this is to obtain test cases and calculate the F-score of the learning algorithm
        xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size=0.2, random_state=0)
        return xTrain, xTest, yTrain, yTest


    # Calculating the F-score
    def calcFScore(xTest, yTest, model, vectorizer):
        
        xTestMatrix = vectorizer.transform(xTest)
        yTestMatrix = np.asarray(yTest)

        result = model.predict(xTestMatrix)
        matrix = confusion_matrix(yTestMatrix, result)

        fScore = f1_score(yTestMatrix, result, pos_label = 0)
        precision = precision_score(yTestMatrix, result, pos_label=0)
        recall = recall_score(yTestMatrix, result, pos_label=0)
        return fScore, precision, recall, matrix

    def predict(msg):
        if lrClassifier.check==0:

            # Create training data
            xTrain, xTest, yTrain, yTest = lrClassifier.store()

            lrClassifier.vectorizer = TfidfVectorizer(stop_words='english', max_df=75)

            yTrainMatrix = np.asarray(yTrain)
            xTrainMatrix = lrClassifier.vectorizer.fit_transform(xTrain)

            # Training LR classifier
            lrClassifier.model.fit(xTrainMatrix, yTrainMatrix)
            fScore, precision, recall, matrix = lrClassifier.calcFScore(xTest, yTest, lrClassifier.model, lrClassifier.vectorizer)
            print("fScore, precision, recall :")
            print(fScore, precision, recall)
            lrClassifier.check=1
            
        return lrClassifier.predict2(msg,lrClassifier.vectorizer,lrClassifier.model)

    # Test new data for Spam
    def predict2(emailBody,vectorizer,model):

        featureMatrix = vectorizer.transform([cleanString(emailBody)])
        result = model.predict(featureMatrix)

        if (1 in result):
            #return "Spam"
            return True
        else:
            #return "Not Spam"
            return False

'''
lrClassifier.predict("This blog is really informative for us. I have read your whole blog you have mentioned points deeply.")
lrClassifier.predict("I also have this problem, ")
lrClassifier.predict("I have a tad issue here about the thorough refining column")
'''
